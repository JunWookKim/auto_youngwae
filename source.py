import openpyxl
import os

#필요없는 열 제거 -> [소속 이름 대분류 중분류], [소속 이름 대분류 중분류] 리스트 생성 ->

wb = openpyxl.load_workbook("test.xlsx")
sheet = wb.worksheets[0]
#print(sheet.cell(2,2).value) #read value
#sheet.delete_cols(5,2)         #E부터 2개 열 제거
#sheet.delete_cols(6)           #F열 제거

#불필요한 열 제거
"""sheet.delete_cols(6,3)
wb.save('./save_test.xlsx')
wb.close()"""

#남아있는 영외활동자 List 생성
L = list()
final_L = list()
"""
for i in range(1,5):
    L.append(sheet.cell(3,i).value)
print(L)

final_L.append(L[:])
print(final_L)

L.clear()
for i in range(1,5):
    L.append(sheet.cell(4,i).value)
final_L.append(L[:])
print(final_L)"""

for i in range(3,sheet.max_row+2):
    final_L.append(L[:])
    L.clear()
    for k in range(1,5):
        L.append(sheet.cell(i,k).value)
print(final_L)
del(final_L[0])
print(final_L)

""" 파일 생성 복사~
wb = openpyxl.Workbook()
wb.save('./save_test.xlsx')
os.system("cp save_test.xlsx cp_save_test.xlsx")
os.system("del save_test.xlsx")
"""